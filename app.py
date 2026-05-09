import streamlit as st
import pandas as pd
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
# MSME Guidelines based on RBI classification
MSME_GUIDELINES = {
    "Micro": {
        "investment_limit": 10000000,  # ₹1 Crore
        "turnover_limit": 50000000,    # ₹5 Crore
        "criteria": "Investment in plant & machinery up to ₹1 Crore",
        "loan_limit": 10000000,
        "features": [
            "Lower interest rates",
            "Easy collateral requirements",
            "Quick loan processing",
            "Eligible for CGTMSE scheme",
            "Government subsidies available"
        ]
    },
    "Small": {
        "investment_limit": 100000000,  # ₹10 Crore
        "turnover_limit": 500000000,    # ₹50 Crore
        "criteria": "Investment in plant & machinery between ₹1-₹10 Crore",
        "loan_limit": 50000000,
        "features": [
            "Moderate interest rates",
            "Moderate collateral requirements",
            "Enhanced loan limits",
            "Access to technology upgradation schemes",
            "Credit guarantee coverage up to 80%"
        ]
    },
    "Medium": {
        "investment_limit": 500000000,  # ₹50 Crore
        "turnover_limit": 2500000000,   # ₹250 Crore
        "criteria": "Investment in plant & machinery between ₹10-₹50 Crore",
        "loan_limit": 100000000,
        "features": [
            "Competitive interest rates",
            "Higher loan limits",
            "Flexible repayment terms",
            "Access to capital markets",
            "Export assistance programs"
        ]
    }
}

def render_guidelines(category):
    """Display guidelines for the identified MSME category"""
    guidelines = MSME_GUIDELINES[category]

    st.subheader(f"📋 Guidelines for {category.upper()} Enterprises")

    col1, col2 = st.columns(2)

    with col1:
        st.metric("Investment Limit", f"₹{guidelines['investment_limit']/100000:,.0f} Lakhs")
        st.metric("Turnover Limit", f"₹{guidelines['turnover_limit']/100000:,.0f} Lakhs")

    with col2:
        st.metric("Maximum Loan Amount", f"₹{guidelines['loan_limit']/100000:,.0f} Lakhs")
        st.write(f"**Definition:** {guidelines['criteria']}")

    st.write("**Key Features & Benefits:**")
    for i, feature in enumerate(guidelines['features'], 1):
        st.write(f"{i}. {feature}")

def assess_cibil_score(score):
    """Assess CIBIL score and provide rating and recommendation"""
    if score is None:
        return "N/A", "No Data", "Commercial CIBIL not available"

    if score >= 750:
        rating = "Excellent"
        category = "A+"
        comment = "Strong creditworthiness. High approval probability."
    elif score >= 700:
        rating = "Very Good"
        category = "A"
        comment = "Good credit history. Likely approval with standard terms."
    elif score >= 650:
        rating = "Good"
        category = "B+"
        comment = "Satisfactory. May require additional documentation or guarantor."
    elif score >= 600:
        rating = "Fair"
        category = "B"
        comment = "Needs improvement. Higher interest rates or lower loan amount may apply."
    elif score >= 550:
        rating = "Poor"
        category = "C"
        comment = "Weak credit profile. Significant scrutiny required. Collateral mandatory."
    else:
        rating = "Very Poor"
        category = "D"
        comment = "Very high risk. Approval unlikely without substantial collateral or guarantor."

    return rating, category, comment

def get_rating_and_comment(param_name, value):
    """Generate rating and comment based on financial parameter value"""
    if param_name == "profit_margin":
        if value >= 15:
            return "A", "Excellent profitability. Strong earnings potential."
        elif value >= 10:
            return "B", "Good profitability. Stable income generation."
        elif value >= 5:
            return "C", "Average profitability. Room for improvement."
        elif value >= 0:
            return "D", "Low profitability. Risk factor for repayment capacity."
        else:
            return "E", "Negative profit. Business is unprofitable. High risk."

    elif param_name == "operating_margin":
        if value >= 20:
            return "A", "Excellent operational efficiency. Costs well controlled."
        elif value >= 15:
            return "B", "Good operational efficiency. Sustainable margins."
        elif value >= 10:
            return "C", "Average efficiency. May face margin pressure."
        elif value >= 5:
            return "D", "Low efficiency. Operational challenges present."
        else:
            return "E", "Very low/negative margins. Operational restructuring needed."

    elif param_name == "roa":
        if value >= 10:
            return "A", "Excellent asset utilization. High returns on assets."
        elif value >= 5:
            return "B", "Good asset utilization and returns."
        elif value >= 2:
            return "C", "Average asset returns. Moderate efficiency."
        elif value >= 0:
            return "D", "Low returns on assets. Asset underutilization."
        else:
            return "E", "Negative ROA. Assets not generating profits."

    elif param_name == "current_ratio":
        if value >= 2.0:
            return "A", "Excellent liquidity. Strong short-term financial health."
        elif value >= 1.5:
            return "B", "Good liquidity position. Can meet short-term obligations."
        elif value >= 1.0:
            return "C", "Acceptable liquidity. Adequate working capital."
        elif value >= 0.8:
            return "D", "Weak liquidity. Potential working capital issues."
        else:
            return "E", "Critical liquidity. Immediate solvency risk."

    elif param_name == "debt_to_equity":
        if value <= 0.5:
            return "A", "Excellent leverage control. Conservative debt levels."
        elif value <= 1.0:
            return "B", "Good debt-equity balance. Sustainable leverage."
        elif value <= 1.5:
            return "C", "Moderate leverage. Acceptable debt levels."
        elif value <= 2.0:
            return "D", "High leverage. Increased financial risk."
        else:
            return "E", "Excessive leverage. Significant solvency concerns."

    return "N/A", "Unable to assess."

def compute_financial_ratios(financials):
    """Calculate important financial ratios with ratings and detailed comments"""
    # Profitability Ratios
    profit_margin = (financials["net_profit"] / financials["gross_revenue"] * 100) if financials["gross_revenue"] > 0 else 0
    operating_margin = ((financials["gross_revenue"] - financials["operating_expenses"]) / financials["gross_revenue"] * 100) if financials["gross_revenue"] > 0 else 0
    roa = (financials["net_profit"] / financials["total_assets"] * 100) if financials["total_assets"] > 0 else 0

    # Liquidity Ratios
    current_ratio = (financials["current_assets"] / financials["current_liabilities"]) if financials["current_liabilities"] > 0 else 0

    # Leverage Ratios
    debt_to_equity = (financials["total_liabilities"] / (financials["total_assets"] - financials["total_liabilities"])) if (financials["total_assets"] - financials["total_liabilities"]) > 0 else 0

    # Get ratings and comments
    npm_rating, npm_comment = get_rating_and_comment("profit_margin", profit_margin)
    om_rating, om_comment = get_rating_and_comment("operating_margin", operating_margin)
    roa_rating, roa_comment = get_rating_and_comment("roa", roa)
    cr_rating, cr_comment = get_rating_and_comment("current_ratio", current_ratio)
    dte_rating, dte_comment = get_rating_and_comment("debt_to_equity", debt_to_equity)

    return {
        "profit_margin": (profit_margin, npm_rating, npm_comment),
        "operating_margin": (operating_margin, om_rating, om_comment),
        "roa": (roa, roa_rating, roa_comment),
        "current_ratio": (current_ratio, cr_rating, cr_comment),
        "debt_to_equity": (debt_to_equity, dte_rating, dte_comment)
    }

def compute_dscr(financials, loan_requirements):
    """Calculate and assess Debt Service Coverage Ratio (DSCR) for term loan"""
    # Calculate Net Operating Income
    net_operating_income = financials["gross_revenue"] - financials["operating_expenses"]

    # Calculate Annual Debt Service for Term Loan
    term_loan = loan_requirements["term_loan"]
    tenure = loan_requirements["tenure_years"]
    interest_rate = loan_requirements["interest_rate"]

    # Annual Principal Repayment (assuming equated annual installments)
    annual_principal = term_loan / tenure

    # Average annual interest (assuming simple interest for approximation)
    average_loan_balance = term_loan / 2
    annual_interest = (average_loan_balance * interest_rate) / 100

    # Total Annual Debt Service
    total_debt_service = annual_principal + annual_interest

    # Calculate DSCR
    dscr = net_operating_income / total_debt_service if total_debt_service > 0 else 0

    # Assess DSCR
    if dscr >= 1.5:
        dscr_rating = "A"
        dscr_comment = "Excellent DSCR. Very strong debt repayment capacity. Low lending risk."
    elif dscr >= 1.25:
        dscr_rating = "B"
        dscr_comment = "Good DSCR. Adequate debt repayment capacity. Acceptable lending risk."
    elif dscr >= 1.0:
        dscr_rating = "C"
        dscr_comment = "Marginal DSCR. Barely adequate to service debt. Moderate risk."
    elif dscr >= 0.8:
        dscr_rating = "D"
        dscr_comment = "Weak DSCR. Limited debt repayment capacity. High lending risk."
    else:
        dscr_rating = "E"
        dscr_comment = "Very Weak DSCR. Insufficient cash flow for debt service. Very high risk."

    return {
        "net_operating_income": net_operating_income,
        "annual_principal": annual_principal,
        "annual_interest": annual_interest,
        "total_debt_service": total_debt_service,
        "dscr": dscr,
        "rating": dscr_rating,
        "comment": dscr_comment
    }

def assess_guarantor_profile(guarantor_details):
    """Assess guarantor profile and provide rating and comment"""
    if not guarantor_details.get("has_guarantor", False):
        return {
            "has_guarantor": False,
            "rating": "None",
            "comment": "No guarantor provided. Higher risk assessment.",
            "support_strength": "Weak"
        }

    cibil_score = guarantor_details.get("cibil_score", 0)
    net_worth = guarantor_details.get("net_worth", 0)

    # Rating based on CIBIL and Net Worth
    if cibil_score >= 750 and net_worth >= 10000000:  # ₹1 Crore
        rating = "Strong"
        comment = "Excellent guarantor profile. Strong financial backing and credit history."
        support_strength = "Strong"
    elif cibil_score >= 700 and net_worth >= 5000000:  # ₹50 Lakhs
        rating = "Good"
        comment = "Good guarantor profile. Adequate financial backing and credit history."
        support_strength = "Good"
    elif cibil_score >= 650 and net_worth >= 2000000:  # ₹20 Lakhs
        rating = "Fair"
        comment = "Fair guarantor profile. Moderate financial backing."
        support_strength = "Fair"
    elif cibil_score >= 600 or net_worth >= 1000000:  # ₹10 Lakhs
        rating = "Weak"
        comment = "Weak guarantor profile. Limited financial backing or credit concerns."
        support_strength = "Weak"
    else:
        rating = "Very Weak"
        comment = "Very weak guarantor profile. Minimal financial backing and credit concerns."
        support_strength = "Very Weak"

    return {
        "has_guarantor": True,
        "rating": rating,
        "comment": comment,
        "support_strength": support_strength
    }

def determine_msme_category(annual_turnover_value):
    """Determine MSME category based on annual turnover"""
    if annual_turnover_value < 50000000:  # Less than ₹5 Crore
        return "Micro"
    elif annual_turnover_value < 500000000:  # Less than ₹50 Crore
        return "Small"
    else:
        return "Medium"

def generate_loan_recommendation(cibil_scores, financial_ratios, dscr_analysis, loan_requirements, annual_turnover, guarantor_analysis):
    """Generate loan recommendation based on all parameters"""
    recommendation_score = 0
    factors = []

    # CIBIL Score Assessment
    if cibil_scores:
        ind_score = cibil_scores["individual"]
        if ind_score >= 750:
            recommendation_score += 25
            factors.append("✓ Excellent individual CIBIL score")
        elif ind_score >= 700:
            recommendation_score += 20
            factors.append("✓ Very good individual CIBIL score")
        elif ind_score >= 650:
            recommendation_score += 15
            factors.append("• Satisfactory individual CIBIL score")
        else:
            factors.append("✗ Weak individual CIBIL score - Risk factor")

    # Financial Ratios Assessment
    if financial_ratios:
        pm_value = financial_ratios['profit_margin'][0]
        cr_value = financial_ratios['current_ratio'][0]
        dte_value = financial_ratios['debt_to_equity'][0]

        if pm_value >= 15:
            recommendation_score += 15
            factors.append("✓ Excellent profit margins")
        elif pm_value >= 10:
            recommendation_score += 10
            factors.append("✓ Good profit margins")
        elif pm_value >= 5:
            recommendation_score += 5
            factors.append("• Average profit margins")
        else:
            factors.append("✗ Low profit margins - Risk factor")

        if cr_value >= 1.5:
            recommendation_score += 15
            factors.append("✓ Strong liquidity position")
        elif cr_value >= 1.0:
            recommendation_score += 10
            factors.append("• Acceptable liquidity")
        else:
            factors.append("✗ Weak liquidity - Working capital concerns")

        if dte_value <= 1.0:
            recommendation_score += 10
            factors.append("✓ Conservative leverage")
        elif dte_value <= 1.5:
            recommendation_score += 5
            factors.append("• Moderate leverage")
        else:
            factors.append("✗ High leverage - Solvency risk")

    # DSCR Assessment
    if dscr_analysis:
        dscr = dscr_analysis['dscr']
        if dscr >= 1.5:
            recommendation_score += 20
            factors.append("✓ Excellent debt service capacity (DSCR >= 1.5)")
        elif dscr >= 1.25:
            recommendation_score += 15
            factors.append("✓ Good debt service capacity (DSCR >= 1.25)")
        elif dscr >= 1.0:
            recommendation_score += 8
            factors.append("• Marginal debt service capacity (DSCR < 1.25)")
        else:
            factors.append("✗ Insufficient debt service capacity - High risk")

    # Guarantor Assessment
    if guarantor_analysis and guarantor_analysis["has_guarantor"]:
        if guarantor_analysis["rating"] == "Strong":
            recommendation_score += 10
            factors.append("✓ Strong guarantor support")
        elif guarantor_analysis["rating"] == "Good":
            recommendation_score += 5
            factors.append("✓ Good guarantor support")
        elif guarantor_analysis["rating"] == "Fair":
            factors.append("• Fair guarantor support")
        else:
            factors.append("✗ Weak guarantor support - Additional risk")

    # Loan to Turnover Assessment
    loan_to_turnover = (loan_requirements['total_loan'] / annual_turnover) * 100
    if loan_to_turnover <= 25:
        recommendation_score += 10
        factors.append("✓ Conservative loan-to-turnover ratio")
    elif loan_to_turnover <= 50:
        recommendation_score += 5
        factors.append("• Moderate loan-to-turnover ratio")
    else:
        factors.append("✗ High loan-to-turnover ratio - Excessive leverage")

    # Generate Recommendation
    if recommendation_score >= 85:
        recommendation = "APPROVED"
        recommendation_text = "Strong recommendation for loan approval. Business demonstrates excellent financial health, strong repayment capacity, and good credit profile."
    elif recommendation_score >= 70:
        recommendation = "CONDITIONAL APPROVAL"
        recommendation_text = "Loan can be approved subject to satisfactory guarantor, higher collateral, or minor conditions."
    elif recommendation_score >= 50:
        recommendation = "NEEDS REVIEW"
        recommendation_text = "Loan requires detailed review. May be approved with additional guarantor, higher interest rate, or lower loan amount."
    else:
        recommendation = "REJECTION RECOMMENDED"
        recommendation_text = "High risk profile. Not recommended for approval unless significant improvements are made or substantial guarantor support provided."

    return {
        "score": recommendation_score,
        "recommendation": recommendation,
        "details": recommendation_text,
        "factors": factors
    }

def generate_excel_proposal(business_name, owner_name, business_type, registration_type, pan_number, gst_number,
                          annual_turnover, loan_purpose, collateral_available, collateral_details,
                          udyam_details, applicant_profile, addresses, guarantor_details, category,
                          loan_requirements, cibil_scores, financial_ratios, dscr_analysis,
                          loan_recommendation, bank_rating, bank_comments, guarantor_analysis,
                          provide_swot, swot_data):
    """Generate Excel proposal with professional formatting"""

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MSME Loan Proposal"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "COMPREHENSIVE MSME LOAN PROPOSAL"
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    # Business Information Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "BUSINESS INFORMATION"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    data = [
        ["Business Name", business_name],
        ["Owner Name", owner_name],
        ["Business Type", business_type],
        ["Registration Type", registration_type],
        ["PAN Number", pan_number],
        ["GST Number", gst_number or "N/A"],
        ["Annual Turnover", f"₹{annual_turnover/100000:,.2f} Lakhs"],
        ["MSME Category", category],
        ["Loan Purpose", loan_purpose],
        ["Collateral Available", collateral_available],
    ]

    if collateral_available.lower() == 'yes':
        data.append(["Collateral Details", collateral_details])

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1  # Empty row

    # Applicant Profile Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "APPLICANT PROFILE"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    data = [
        ["Age", applicant_profile.get('age', 'N/A')],
        ["Education", applicant_profile.get('education', 'N/A')],
        ["Business Experience", f"{applicant_profile.get('business_experience', 'N/A')} years"],
        ["Previous Loan Experience", "Yes" if applicant_profile.get('previous_loans') else "No"],
    ]

    if applicant_profile.get('previous_loan_details'):
        data.append(["Previous Loan Details", applicant_profile['previous_loan_details']])

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # Udyam Registration Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "UDYAM REGISTRATION"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    if udyam_details.get('registered'):
        data = [
            ["Udyam Registered", "Yes"],
            ["Udyam Number", udyam_details.get('udyam_number', 'N/A')],
            ["Registration Date", udyam_details.get('registration_date', 'N/A')],
        ]
    else:
        data = [["Udyam Registered", "No"]]

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # Address Details Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "ADDRESS DETAILS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    firm_addr = addresses['firm_address']
    promoter_addr = addresses['promoter_address']

    data = [
        ["Firm Address", f"{firm_addr['street']}, {firm_addr['city']}, {firm_addr['state']} - {firm_addr['pincode']}"],
        ["Promoter Address", f"{promoter_addr['street']}, {promoter_addr['city']}, {promoter_addr['state']} - {promoter_addr['pincode']}"],
    ]

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # Guarantor Details Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "GUARANTOR DETAILS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    if guarantor_details.get('has_guarantor'):
        data = [
            ["Guarantor Name", guarantor_details.get('name', 'N/A')],
            ["Relationship", guarantor_details.get('relationship', 'N/A')],
            ["Contact", guarantor_details.get('contact', 'N/A')],
            ["PAN", guarantor_details.get('pan', 'N/A')],
            ["Net Worth", f"₹{guarantor_details.get('net_worth', 0)/100000:,.2f} Lakhs"],
            ["CIBIL Score", guarantor_details.get('cibil_score', 'N/A')],
        ]
        if guarantor_analysis:
            data.append(["Guarantor Rating", guarantor_analysis.get('rating', 'N/A')])
            data.append(["Support Strength", guarantor_analysis.get('support_strength', 'N/A')])
    else:
        data = [["Guarantor", "No Guarantor Provided"]]

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # Loan Requirements Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "LOAN REQUIREMENTS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    data = [
        ["Working Capital Loan", f"₹{loan_requirements['working_capital_loan']/100000:,.2f} Lakhs"],
        ["Term Loan", f"₹{loan_requirements['term_loan']/100000:,.2f} Lakhs"],
        ["Total Loan Required", f"₹{loan_requirements['total_loan']/100000:,.2f} Lakhs"],
        ["Loan-to-Turnover Ratio", f"{(loan_requirements['total_loan']/annual_turnover)*100:.2f}%"],
    ]

    if loan_requirements["term_loan"] > 0:
        data.extend([
            ["Term Loan Tenure", f"{loan_requirements['tenure_years']} years"],
            ["Expected Interest Rate", f"{loan_requirements['interest_rate']:.2f}% p.a."],
        ])

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # CIBIL Scores Section (if provided)
    if cibil_scores:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "CREDIT ASSESSMENT (CIBIL)"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        data = [
            ["Individual/Proprietor CIBIL Score", cibil_scores["individual"]],
        ]

        if cibil_scores["commercial"]:
            data.append(["Commercial/Business CIBIL Score", cibil_scores["commercial"]])

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = normal_font
            row += 1

        row += 1

    # Financial Ratios Section (if provided)
    if financial_ratios:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "FINANCIAL ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        data = [
            ["Net Profit Margin", f"{financial_ratios['profit_margin'][0]:.2f}%"],
            ["Operating Margin", f"{financial_ratios['operating_margin'][0]:.2f}%"],
            ["Return on Assets (ROA)", f"{financial_ratios['roa'][0]:.2f}%"],
            ["Current Ratio", f"{financial_ratios['current_ratio'][0]:.2f}"],
            ["Debt-to-Equity Ratio", f"{financial_ratios['debt_to_equity'][0]:.2f}"],
        ]

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = normal_font
            row += 1

        row += 1

    # DSCR Analysis Section (if provided)
    if dscr_analysis:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "DEBT SERVICE COVERAGE RATIO (DSCR)"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        data = [
            ["Net Operating Income", f"₹{dscr_analysis['net_operating_income']/100000:,.2f} Lakhs"],
            ["Annual Debt Service", f"₹{dscr_analysis['total_debt_service']/100000:,.2f} Lakhs"],
            ["DSCR", f"{dscr_analysis['dscr']:.2f}"],
            ["DSCR Rating", dscr_analysis['rating']],
        ]

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = normal_font
            row += 1

        row += 1

    # Loan Recommendation Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "LOAN RECOMMENDATION"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    data = [
        ["Recommendation Score", f"{loan_recommendation['score']}/100"],
        ["Overall Recommendation", loan_recommendation['recommendation']],
        ["Details", loan_recommendation['details']],
    ]

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # Bank Internal Assessment Section
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "BANK INTERNAL ASSESSMENT"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    data = [
        ["Bank Internal Rating", bank_rating],
        ["Internal Comments", bank_comments],
    ]

    for label, value in data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = normal_font
        row += 1

    row += 1

    # SWOT Analysis Section (if provided)
    if provide_swot and swot_data:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "SWOT ANALYSIS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Strengths
        if swot_data.get('strengths'):
            ws[f'A{row}'] = "Strengths:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for strength in swot_data['strengths']:
                ws[f'B{row}'] = f"• {strength}"
                ws[f'B{row}'].font = normal_font
                row += 1

        # Weaknesses
        if swot_data.get('weaknesses'):
            ws[f'A{row}'] = "Weaknesses:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for weakness in swot_data['weaknesses']:
                ws[f'B{row}'] = f"• {weakness}"
                ws[f'B{row}'].font = normal_font
                row += 1

        # Opportunities
        if swot_data.get('opportunities'):
            ws[f'A{row}'] = "Opportunities:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for opportunity in swot_data['opportunities']:
                ws[f'B{row}'] = f"• {opportunity}"
                ws[f'B{row}'].font = normal_font
                row += 1

        # Threats
        if swot_data.get('threats'):
            ws[f'A{row}'] = "Threats:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for threat in swot_data['threats']:
                ws[f'B{row}'] = f"• {threat}"
                ws[f'B{row}'].font = normal_font
                row += 1

    # Apply borders to all cells
    for row_num in range(1, row):
        for col_num in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def main():
    st.set_page_config(
        page_title="MSME Loan Proposal Generator",
        page_icon="🏦",
        layout="wide"
    )

    st.title("🏦 Comprehensive MSME Loan Proposal Generator")
    st.markdown("Generate professional loan proposals for Indian banks with automated analysis and Excel export.")

    # Initialize session state
    if 'proposal_generated' not in st.session_state:
        st.session_state.proposal_generated = False

    with st.form("loan_proposal_form"):
        st.header("📝 Business Information")

        col1, col2 = st.columns(2)

        with col1:
            business_name = st.text_input("Business Name *", placeholder="Enter business name")
            owner_name = st.text_input("Owner/Proprietor Name *", placeholder="Enter owner name")
            business_type = st.selectbox("Business Type *",
                                       ["Manufacturing", "Services", "Trading", "Retail", "Construction", "Other"])
            registration_type = st.selectbox("Registration Type *",
                                           ["Proprietorship", "Partnership", "Private Limited", "LLP", "Other"])

        with col2:
            pan_number = st.text_input("PAN Number *", placeholder="Enter PAN number")
            gst_number = st.text_input("GST Number", placeholder="Enter GST number (if applicable)")
            annual_turnover = st.number_input("Annual Turnover (in Lakhs) *", min_value=0.0, step=0.01)
            loan_purpose = st.selectbox("Purpose of Loan *",
                                      ["Working Capital", "Term Loan", "Working Capital + Term Loan", "Business Expansion"])

        # MSME Category determination
        annual_turnover_value = annual_turnover * 100000
        category = determine_msme_category(annual_turnover_value)

        st.info(f"📊 **MSME Category:** {category} (Based on ₹{annual_turnover:,.2f} Lakhs turnover)")

        # Collateral Information
        collateral_available = st.radio("Collateral Available? *", ["Yes", "No"], horizontal=True)
        collateral_details = ""
        if collateral_available == "Yes":
            collateral_details = st.text_area("Collateral Details", placeholder="Describe available collateral")

        st.header("👤 Applicant Profile")

        col1, col2 = st.columns(2)

        with col1:
            age = st.text_input("Age", placeholder="Enter age")
            education = st.selectbox("Educational Qualification",
                                   ["Below 10th", "10th Pass", "12th Pass", "Graduate", "Post Graduate", "Professional", "Other"])

        with col2:
            business_experience = st.number_input("Years of Business Experience", min_value=0, max_value=50, step=1)
            previous_loans = st.radio("Previous Loan Experience?", ["Yes", "No"], horizontal=True)
            previous_loan_details = ""
            if previous_loans == "Yes":
                previous_loan_details = st.text_area("Previous Loan Details",
                                                   placeholder="Bank name, amount, status, etc.")

        st.header("📋 Registration Details")

        udyam_registered = st.radio("Is business Udyam registered? *", ["Yes", "No"], horizontal=True)
        udyam_number = ""
        registration_date = ""
        if udyam_registered == "Yes":
            col1, col2 = st.columns(2)
            with col1:
                udyam_number = st.text_input("Udyam Registration Number")
            with col2:
                registration_date = st.text_input("Registration Date (DD/MM/YYYY)")

        st.header("🏠 Address Details")

        st.subheader("Firm/Business Address")
        col1, col2 = st.columns(2)
        with col1:
            firm_street = st.text_input("Street Address (Firm)")
            firm_city = st.text_input("City (Firm)")
        with col2:
            firm_state = st.text_input("State (Firm)")
            firm_pincode = st.text_input("Pincode (Firm)")

        st.subheader("Promoter/Owner Residential Address")
        col1, col2 = st.columns(2)
        with col1:
            promoter_street = st.text_input("Street Address (Promoter)")
            promoter_city = st.text_input("City (Promoter)")
        with col2:
            promoter_state = st.text_input("State (Promoter)")
            promoter_pincode = st.text_input("Pincode (Promoter)")

        st.header("🤝 Guarantor Details")

        has_guarantor = st.radio("Is there a guarantor? *", ["Yes", "No"], horizontal=True)

        guarantor_name = ""
        guarantor_relationship = ""
        guarantor_contact = ""
        guarantor_pan = ""
        guarantor_net_worth = 0.0
        guarantor_cibil = 0

        if has_guarantor == "Yes":
            col1, col2 = st.columns(2)
            with col1:
                guarantor_name = st.text_input("Guarantor Name")
                guarantor_relationship = st.text_input("Relationship with Applicant")
                guarantor_contact = st.text_input("Guarantor Contact Number")
            with col2:
                guarantor_pan = st.text_input("Guarantor PAN")
                guarantor_net_worth = st.number_input("Guarantor Net Worth (in Lakhs)", min_value=0.0, step=0.01)
                guarantor_cibil = st.number_input("Guarantor CIBIL Score", min_value=300, max_value=900, step=1)

        st.header("💰 Loan Requirements")

        col1, col2 = st.columns(2)

        with col1:
            working_capital_loan = st.number_input("Working Capital Loan Required (in Lakhs)", min_value=0.0, step=0.01)
            term_loan = st.number_input("Term Loan Required (in Lakhs)", min_value=0.0, step=0.01)

        with col2:
            tenure_years = st.number_input("Term Loan Tenure (in years)", min_value=1, max_value=30, step=1)
            interest_rate = st.number_input("Expected Interest Rate (% per annum)", min_value=0.0, max_value=20.0, step=0.1)

        st.header("📊 Optional Analysis")

        provide_cibil = st.checkbox("Provide CIBIL Score Information")
        individual_cibil = 0
        commercial_cibil = 0

        if provide_cibil:
            col1, col2 = st.columns(2)
            with col1:
                individual_cibil = st.number_input("Individual/Proprietor CIBIL Score", min_value=300, max_value=900, step=1)
            with col2:
                commercial_cibil = st.number_input("Commercial/Business CIBIL Score (0 if not available)", min_value=0, max_value=900, step=1)

        provide_financials = st.checkbox("Provide Financial Information")
        gross_revenue = operating_expenses = net_profit = 0.0
        current_assets = current_liabilities = total_assets = total_liabilities = 0.0

        if provide_financials:
            st.subheader("Financial Data (in Lakhs)")
            col1, col2 = st.columns(2)
            with col1:
                gross_revenue = st.number_input("Gross Annual Revenue", min_value=0.0, step=0.01)
                operating_expenses = st.number_input("Annual Operating Expenses", min_value=0.0, step=0.01)
                net_profit = st.number_input("Net Profit", min_value=0.0, step=0.01)
            with col2:
                current_assets = st.number_input("Current Assets", min_value=0.0, step=0.01)
                current_liabilities = st.number_input("Current Liabilities", min_value=0.0, step=0.01)
                total_assets = st.number_input("Total Assets", min_value=0.0, step=0.01)
                total_liabilities = st.number_input("Total Liabilities", min_value=0.0, step=0.01)

        provide_swot = st.checkbox("Provide SWOT Analysis")
        strengths = weaknesses = opportunities = threats = ""

        if provide_swot:
            st.subheader("SWOT Analysis (comma-separated)")
            col1, col2 = st.columns(2)
            with col1:
                strengths = st.text_area("Strengths", placeholder="Enter strengths separated by commas")
                weaknesses = st.text_area("Weaknesses", placeholder="Enter weaknesses separated by commas")
            with col2:
                opportunities = st.text_area("Opportunities", placeholder="Enter opportunities separated by commas")
                threats = st.text_area("Threats", placeholder="Enter threats separated by commas")

        st.header("🏦 Bank Assessment")

        bank_rating = st.text_input("Bank Internal Rating (e.g., A1, B2, C1, D2...)", placeholder="Enter alphanumeric rating")
        bank_comments = st.text_area("Internal Comments/Remarks", placeholder="Enter bank assessment comments")

        # Submit button
        submitted = st.form_submit_button("🚀 Generate Proposal", type="primary")

        if submitted:
            # Validate required fields
            if not business_name or not owner_name or not pan_number or annual_turnover <= 0:
                st.error("Please fill in all required fields marked with *")
                return

            # Process data
            with st.spinner("Generating comprehensive loan proposal..."):

                # Prepare data structures
                udyam_details = {
                    "registered": udyam_registered == "Yes",
                    "udyam_number": udyam_number,
                    "registration_date": registration_date
                }

                applicant_profile = {
                    "age": age,
                    "education": education,
                    "business_experience": business_experience,
                    "previous_loans": previous_loans == "Yes",
                    "previous_loan_details": previous_loan_details
                }

                addresses = {
                    "firm_address": {
                        "street": firm_street,
                        "city": firm_city,
                        "state": firm_state,
                        "pincode": firm_pincode
                    },
                    "promoter_address": {
                        "street": promoter_street,
                        "city": promoter_city,
                        "state": promoter_state,
                        "pincode": promoter_pincode
                    }
                }

                guarantor_details = {
                    "has_guarantor": has_guarantor == "Yes",
                    "name": guarantor_name,
                    "relationship": guarantor_relationship,
                    "contact": guarantor_contact,
                    "pan": guarantor_pan,
                    "net_worth": guarantor_net_worth * 100000,
                    "cibil_score": guarantor_cibil
                }

                loan_requirements = {
                    "working_capital_loan": working_capital_loan * 100000,
                    "term_loan": term_loan * 100000,
                    "tenure_years": tenure_years,
                    "interest_rate": interest_rate,
                    "total_loan": (working_capital_loan + term_loan) * 100000,
                }

                # Optional data
                cibil_scores = None
                if provide_cibil:
                    cibil_scores = {
                        "individual": individual_cibil,
                        "commercial": commercial_cibil if commercial_cibil > 0 else None
                    }

                financials = None
                financial_ratios = None
                dscr_analysis = None
                if provide_financials:
                    financials = {
                        "gross_revenue": gross_revenue * 100000,
                        "operating_expenses": operating_expenses * 100000,
                        "net_profit": net_profit * 100000,
                        "current_assets": current_assets * 100000,
                        "current_liabilities": current_liabilities * 100000,
                        "total_assets": total_assets * 100000,
                        "total_liabilities": total_liabilities * 100000,
                    }
                    financial_ratios = compute_financial_ratios(financials)
                    if term_loan > 0:
                        dscr_analysis = compute_dscr(financials, loan_requirements)

                # Construct SWOT data
                swot_data = None
                if provide_swot:
                    swot_data = {
                        "strengths": [s.strip() for s in strengths.split(',') if s.strip()],
                        "weaknesses": [w.strip() for w in weaknesses.split(',') if w.strip()],
                        "opportunities": [o.strip() for o in opportunities.split(',') if o.strip()],
                        "threats": [t.strip() for t in threats.split(',') if t.strip()],
                    }

                # Guarantor analysis
                guarantor_analysis = assess_guarantor_profile(guarantor_details)

                # Generate recommendation
                loan_recommendation = generate_loan_recommendation(
                    cibil_scores, financial_ratios, dscr_analysis, loan_requirements,
                    annual_turnover_value, guarantor_analysis
                )

                # Generate Excel
                excel_file = generate_excel_proposal(
                    business_name, owner_name, business_type, registration_type, pan_number, gst_number,
                    annual_turnover_value, loan_purpose, collateral_available, collateral_details,
                    udyam_details, applicant_profile, addresses, guarantor_details, category,
                    loan_requirements, cibil_scores, financial_ratios, dscr_analysis,
                    loan_recommendation, bank_rating, bank_comments, guarantor_analysis,
                    provide_swot, swot_data
                )

                # Store results in session state
                st.session_state.proposal_generated = True
                st.session_state.results = {
                    'business_name': business_name,
                    'owner_name': owner_name,
                    'category': category,
                    'loan_recommendation': loan_recommendation,
                    'guarantor_analysis': guarantor_analysis,
                    'financial_ratios': financial_ratios,
                    'dscr_analysis': dscr_analysis,
                    'cibil_scores': cibil_scores,
                    'excel_file': excel_file,
                    'provide_swot': provide_swot,
                    'swot_data': swot_data
                }

                st.success("✅ Loan proposal generated successfully!")
                st.rerun()

    # Display results if generated
    if st.session_state.proposal_generated:
        results = st.session_state.results

        st.header("📊 Analysis Results")

        # MSME Guidelines
        with st.expander("📋 MSME Guidelines", expanded=True):
            render_guidelines(results['category'])

        # Loan Recommendation
        with st.expander("🎯 Loan Recommendation", expanded=True):
            rec = results['loan_recommendation']
            col1, col2 = st.columns(2)

            with col1:
                st.metric("Recommendation Score", f"{rec['score']}/100")
                st.write(f"**Overall Recommendation:** {rec['recommendation']}")

            with col2:
                if rec['score'] >= 85:
                    st.success("✅ Strong approval recommendation")
                elif rec['score'] >= 70:
                    st.warning("⚠️ Conditional approval")
                elif rec['score'] >= 50:
                    st.info("🔍 Needs detailed review")
                else:
                    st.error("❌ High risk - Rejection recommended")

            st.write(f"**Details:** {rec['details']}")

            st.write("**Analysis Factors:**")
            for factor in rec['factors']:
                st.write(f"• {factor}")

        # Guarantor Analysis
        if results['guarantor_analysis']['has_guarantor']:
            with st.expander("🤝 Guarantor Analysis", expanded=True):
                ga = results['guarantor_analysis']
                st.write(f"**Rating:** {ga['rating']}")
                st.write(f"**Support Strength:** {ga['support_strength']}")
                st.write(f"**Comment:** {ga['comment']}")

        # Financial Analysis
        if results['financial_ratios']:
            with st.expander("📈 Financial Analysis", expanded=True):
                fr = results['financial_ratios']
                col1, col2 = st.columns(2)

                with col1:
                    st.metric("Net Profit Margin", f"{fr['profit_margin'][0]:.2f}%")
                    st.metric("Operating Margin", f"{fr['operating_margin'][0]:.2f}%")
                    st.metric("Return on Assets", f"{fr['roa'][0]:.2f}%")

                with col2:
                    st.metric("Current Ratio", f"{fr['current_ratio'][0]:.2f}")
                    st.metric("Debt-to-Equity", f"{fr['debt_to_equity'][0]:.2f}")

                st.write("**Detailed Comments:**")
                st.write(f"• Profit Margin: {fr['profit_margin'][2]}")
                st.write(f"• Operating Margin: {fr['operating_margin'][2]}")
                st.write(f"• ROA: {fr['roa'][2]}")
                st.write(f"• Current Ratio: {fr['current_ratio'][2]}")
                st.write(f"• Debt-to-Equity: {fr['debt_to_equity'][2]}")

        # DSCR Analysis
        if results['dscr_analysis']:
            with st.expander("💰 DSCR Analysis", expanded=True):
                dscr = results['dscr_analysis']
                col1, col2 = st.columns(2)

                with col1:
                    st.metric("DSCR", f"{dscr['dscr']:.2f}")
                    st.metric("Net Operating Income", f"₹{dscr['net_operating_income']/100000:,.2f} Lakhs")

                with col2:
                    st.metric("Annual Debt Service", f"₹{dscr['total_debt_service']/100000:,.2f} Lakhs")
                    st.write(f"**Rating:** {dscr['rating']}")

                st.write(f"**Comment:** {dscr['comment']}")

        # CIBIL Assessment
        if results['cibil_scores']:
            with st.expander("📊 CIBIL Assessment", expanded=True):
                cibil = results['cibil_scores']
                ind_rating, ind_category, ind_comment = assess_cibil_score(cibil["individual"])

                st.metric("Individual CIBIL Score", cibil["individual"])
                st.write(f"**Category:** {ind_category}")
                st.write(f"**Assessment:** {ind_comment}")

                if cibil["commercial"]:
                    com_rating, com_category, com_comment = assess_cibil_score(cibil["commercial"])
                    st.metric("Commercial CIBIL Score", cibil["commercial"])
                    st.write(f"**Category:** {com_category}")
                    st.write(f"**Assessment:** {com_comment}")

        # SWOT Analysis
        if results['provide_swot'] and results['swot_data']:
            with st.expander("🎯 SWOT Analysis", expanded=True):
                swot = results['swot_data']
                col1, col2 = st.columns(2)

                with col1:
                    st.subheader("✅ Strengths")
                    for s in swot.get('strengths', []):
                        st.write(f"• {s}")

                    st.subheader("❌ Weaknesses")
                    for w in swot.get('weaknesses', []):
                        st.write(f"• {w}")

                with col2:
                    st.subheader("🚀 Opportunities")
                    for o in swot.get('opportunities', []):
                        st.write(f"• {o}")

                    st.subheader("⚠️ Threats")
                    for t in swot.get('threats', []):
                        st.write(f"• {t}")

        # Download Excel
        st.header("📥 Download Proposal")
        st.download_button(
            label="📊 Download Complete Excel Proposal",
            data=results['excel_file'],
            file_name=f"MSME_Loan_Proposal_{results['business_name'].replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Important Notes
        with st.expander("📋 Important Notes", expanded=False):
            st.markdown("""
            **Important Notes:**
            • This is a comprehensive proposal template for Indian bank loan applications
            • Ensure all information is accurate and verified
            • All financial data should be supported by audited documents
            • MSME category determines eligibility under government schemes
            • DSCR minimum of 1.25x is typically required for term loan approval
            • Guarantor support strengthens the application
            • Udyam registration enhances eligibility for government benefits
            """)

if __name__ == "__main__":
    main()